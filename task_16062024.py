# Написать программу «успеваемость». Пользователь
# вводит 10 оценок студента. Оценки от 1 до 12. Реализовать
# меню для пользователя:
# ■ Вывод оценок (вывод содержимого списка);
# ■ Пересдача экзамена (пользователь вводит номер элемента списка и новую оценку);
# ■ Выходит ли стипендия (стипендия выходит, если
# средний бал не ниже 10.7);
# ■ Вывод отсортированного списка оценок: по возрастанию или убыванию.

import os
import statistics
import openpyxl
from colorama import init, Fore, Style
init()  # Инициализация colorama


def print_colorful(text, color=Fore.WHITE):
    print(color + text + Style.RESET_ALL)


def get_data(path: str) -> list:
    data = []
    wb = openpyxl.load_workbook(path)
    sheet = wb["Лист1"]
    for i in range(2, sheet.max_row + 1):
        marks = []
        for j in range(2, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value is None:
                temp = 0
            else:
                try:
                    temp = int(sheet.cell(row=i, column=j).value)
                except ValueError:
                    temp = 0
            marks.append(0 if temp < 0 or temp > 12 else temp)
        data.append({"fio": sheet[f"A{i}"].value, "marks": marks})
    return data


def set_data(path: str, data: list) -> None:
    wb = openpyxl.load_workbook(path)
    sheet = wb["Лист1"]
    for i in range(2, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            sheet.cell(row=i, column=j).value = ""
    for i in range(len(data)):
        sheet.cell(row=i+2, column=1).value = data[i]['fio']
        for j in range(len(data[i]["marks"])):
            sheet.cell(row=i+2, column=j+2).value = data[i]["marks"][j]
    wb.save(path)
    wb.close()


def print_data(data: list) -> None:
    for i in range(len(data)):
        print(f"{i+1}. {data[i]['fio']}, оценки: {data[i]["marks"]}, средняя оценка {statistics.mean(data[i]["marks"])} - {"положена стипендия" if statistics.mean(data[i]["marks"]) >= 10.7 else "не положена стипендия"}")


def print_sorted(ls: list, to_up=True) -> list:
    for i in range(1, len(ls)):
        item_to_insert = ls[i]
        j = i - 1
        while j >= 0 and not to_up and statistics.mean(item_to_insert["marks"]) > statistics.mean(ls[j]["marks"]) or j >= 0 and to_up and statistics.mean(item_to_insert["marks"]) < statistics.mean(ls[j]["marks"]):
            ls[j + 1] = ls[j]
            j = j - 1
        ls[j+1] = item_to_insert
    print_data(ls)
    return ls


data_group1 = []
print("Программа \"Анализ успеваемости струдентов\"")
while True:
    print_colorful("Меню:\n1 - загрузка данных из файла и вывод оценок с данными по стипендии,\n2 - поиск студента по номеру в списке для корректировки оценки за экзамен,\n3 - запись скорректированного массива данных в файл,\n4 - вывод отсортированного списка по средней оценке по возрастанию,\n5 - вывод отсортированного списка по средней оценке по убыванию,\n6 - завершение программы.", Fore.LIGHTMAGENTA_EX)
    try:
        selected_option = int(input("Введите пункт меню: "))
    except ValueError:
        print("Ошибка ввода при выборе пункта меню.")
        continue
    match selected_option:
        case 1:
            try:
                data_group1 = get_data(os.getcwd() + "\\Exam_16062024.xlsx")
                print_data(data_group1)
            except Exception:
                print(f"Ошибка при зазрузке данных из файла {os.getcwd() + "\\Exam_16062024.xlsx"}. Проверьте наличие файла и его формат.")
        case 2:
            if len(data_group1) == 0:
                print("Сначала выберите пункт 1, в настоящее время данные не загружены.")
                continue
            print(f"В программу загружены сведения по студентам в количестве {len(data_group1)}. Выберите номер студента в соответствующем диапазоне.")
            try:
                selected_student = int(input("Введите номер студента: "))
            except ValueError:
                print("Ошибка ввода данных.")
                continue
            if selected_student < 1 or selected_student > len(data_group1):
                print("Ошибка ввода данных, номер студента вне допустимого диапазона.")
            else:
                try:
                    exam_mark = int(input("Введите оценку по итогам пересдачи экзамена (0 - 12): "))
                except ValueError:
                    print("Ошибка ввода данных.")
                    continue
                if exam_mark < 0 or exam_mark > 12:
                    print("Ошибка ввода данных, оценка вне допустимого диапазона.")
                    continue
                data_group1[selected_student-1]["marks"][len(data_group1[selected_student-1]["marks"])-1] = exam_mark
        case 3:
            if len(data_group1) == 0:
                print("Сначала выберите пункт 1, в настоящее время данные не загружены.")
                continue
            set_data(os.getcwd() + "\\Exam_16062024.xlsx", data_group1)
        case 4:
            if len(data_group1) == 0:
                print("Сначала выберите пункт 1, в настоящее время данные не загружены.")
                continue
            data_group1 = print_sorted(data_group1, to_up=True)
        case 5:
            if len(data_group1) == 0:
                print("Сначала выберите пункт 1, в настоящее время данные не загружены.")
                continue
            data_group1 = print_sorted(data_group1, to_up=False)
        case 6:
            print("Программа завершает свою работу...")
            break
        case _:
            print("Введенный номер пункта меню вне допустимого диапазона.")
            continue
